"""
VPL Exhibition Enquiry Form
----------------------------
Flask + SQLite web app to capture exhibition leads, store them reliably
(multi-device, concurrent writes), and export to CSV / Excel.

Product Interested In, Action, and VPL Coordinator are all DB-backed
option lists managed from the Master Settings page (/settings) - add or
remove options there and the form updates immediately, no code changes.

Run locally:
    pip install -r requirements.txt
    python app.py
    -> open http://127.0.0.1:5000 on this machine
    -> open http://<this-machine-LAN-IP>:5000 on other devices on the
       same network (the app prints this IP on startup)

Deployment: see README.md for AWS EC2 + gunicorn + systemd instructions.
"""

import csv
import io
import json
import os
import socket
import sqlite3
import uuid
from datetime import datetime, timedelta
from functools import wraps

from flask import (
    Flask, render_template, request, jsonify, g,
    Response, send_file, redirect, url_for, flash, session
)
from werkzeug.security import generate_password_hash, check_password_hash

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "leads.db")

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("VPL_SECRET_KEY", "vpl-exhibition-secret-change-me")

# No session timeout during the exhibition: once logged in, stay logged in
# for 30 days (until you explicitly log out), instead of the browser's
# default "expires when the browser closes" behavior.
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)
app.config["SESSION_REFRESH_EACH_REQUEST"] = True

# These protect only the Master Settings page (adding/removing dropdown
# options) - they do NOT block anyone from filling the enquiry form or
# viewing leads. Used only to seed the DB the very first time the app
# runs; after that, credentials live in the database and can be changed
# from the Master Settings page itself.
DEFAULT_ADMIN_USERNAME = os.environ.get("VPL_ADMIN_USERNAME", "admin")
DEFAULT_ADMIN_PASSWORD = os.environ.get("VPL_ADMIN_PASSWORD", "123123")

# field_key values used across the app - keep this as the single source of
# truth so typos can't silently create a orphan/unused list.
FIELD_PRODUCT_TYPES = "product_types"
FIELD_ACTIONS = "actions"
FIELD_COORDINATOR = "vpl_coordinator"
FIELD_COUNTRY_CODE = "country_codes"

# used as the default pre-selected option in the Country Code dropdown
DEFAULT_COUNTRY_CODE_VALUE = "+91 India"

DEFAULT_OPTIONS = {
    FIELD_PRODUCT_TYPES: ["Mono Carton", "Corrugated Box", "3 Ply", "Shipper Box", "Rigid Box"],
    FIELD_ACTIONS: ["Rate", "Sample", "KLD", "Option", "Visit after Exhibition", "E-meet after Exhibition"],
    FIELD_COORDINATOR: [],  # left empty on purpose - add your team's names via Master Settings
    FIELD_COUNTRY_CODE: [
        DEFAULT_COUNTRY_CODE_VALUE, "+1 USA/Canada", "+44 UK", "+971 UAE", "+966 Saudi Arabia",
        "+974 Qatar", "+968 Oman", "+965 Kuwait", "+973 Bahrain", "+61 Australia",
        "+65 Singapore", "+60 Malaysia", "+66 Thailand", "+62 Indonesia", "+63 Philippines",
        "+86 China", "+81 Japan", "+82 South Korea", "+92 Pakistan", "+880 Bangladesh",
        "+94 Sri Lanka", "+977 Nepal", "+95 Myanmar", "+49 Germany", "+33 France",
        "+39 Italy", "+34 Spain", "+31 Netherlands", "+46 Sweden", "+41 Switzerland",
        "+7 Russia", "+27 South Africa", "+234 Nigeria", "+254 Kenya", "+20 Egypt",
        "+55 Brazil", "+52 Mexico", "+54 Argentina", "+64 New Zealand", "+353 Ireland",
    ],
}

FIELD_LABELS = {
    FIELD_PRODUCT_TYPES: "Product Interested In",
    FIELD_ACTIONS: "Action",
    FIELD_COORDINATOR: "VPL Coordinator",
    FIELD_COUNTRY_CODE: "Country Code",
}

EXPORT_HEADERS = [
    "Customer Name", "Contact No.", "Date", "Source",
    "Product Interested In", "Product Description",
    "Dimensions (L x W x H)", "Material/GSM", "Printing & Finishing",
    "Estimated Quantity", "Required By", "Requirement / Remarks",
    "Reference Carton", "Action", "VPL Coordinator", "Remark", "Saved At"
]

EXPORT_FIELD_ORDER = [
    "name", "contact", "entry_date", "source",
    "product_types", "product_description",
    "dimensions", "material_gsm", "printing_finishing",
    "est_quantity", "required_by", "remarks",
    "reference_carton", "actions", "vpl_coordinator", "final_remark", "created_at"
]


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_db():
    if "db" not in g:
        conn = sqlite3.connect(DB_PATH, timeout=15)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON;")
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    """Create tables (if missing), enable WAL mode for concurrency, and
    seed default dropdown options only on a brand new database."""
    conn = sqlite3.connect(DB_PATH, timeout=15)
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")

        conn.execute("""
            CREATE TABLE IF NOT EXISTS leads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                uid TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                contact TEXT NOT NULL,
                entry_date TEXT,
                source TEXT,
                product_types TEXT,
                product_description TEXT,
                dimensions TEXT,
                material_gsm TEXT,
                printing_finishing TEXT,
                est_quantity TEXT,
                required_by TEXT,
                remarks TEXT,
                reference_carton TEXT,
                actions TEXT,
                vpl_coordinator TEXT,
                final_remark TEXT,
                created_at TEXT NOT NULL
            );
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_leads_created_at ON leads(created_at);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_leads_name ON leads(name);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_leads_contact ON leads(contact);")

        # migration: add custom_data column if this is an existing DB from
        # before custom fields existed (CREATE TABLE IF NOT EXISTS above
        # won't add new columns to an already-existing leads table)
        existing_cols = [row[1] for row in conn.execute("PRAGMA table_info(leads)").fetchall()]
        if "custom_data" not in existing_cols:
            conn.execute("ALTER TABLE leads ADD COLUMN custom_data TEXT DEFAULT '{}';")

        conn.execute("""
            CREATE TABLE IF NOT EXISTS custom_fields (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                label TEXT UNIQUE NOT NULL COLLATE NOCASE,
                created_at TEXT NOT NULL
            );
        """)

        # migration: add sort_order if this is an existing DB from before
        # Date/Source were folded into the custom fields list
        cf_cols = [row[1] for row in conn.execute("PRAGMA table_info(custom_fields)").fetchall()]
        if "sort_order" not in cf_cols:
            conn.execute("ALTER TABLE custom_fields ADD COLUMN sort_order INTEGER;")
            # preserve the relative order fields were originally added in,
            # offset so Date/Source (order 0 and 1, pinned below) stay on top
            conn.execute("UPDATE custom_fields SET sort_order = id + 100 WHERE sort_order IS NULL;")

        # Date and Source used to be fixed, hardcoded fields. They now live
        # in this same removable/addable list, pinned to the top (sort_order
        # 0 and 1) whether this is a brand new DB or an upgrade from an
        # older one that already has other custom fields.
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for pinned_label, pinned_order in (("Date", 0), ("Source", 1)):
            existing_row = conn.execute(
                "SELECT id FROM custom_fields WHERE label = ? COLLATE NOCASE", (pinned_label,)
            ).fetchone()
            if existing_row:
                conn.execute(
                    "UPDATE custom_fields SET sort_order = ? WHERE id = ?",
                    (pinned_order, existing_row[0]),
                )
            else:
                conn.execute(
                    "INSERT INTO custom_fields (label, created_at, sort_order) VALUES (?, ?, ?)",
                    (pinned_label, now, pinned_order),
                )

        conn.execute("""
            CREATE TABLE IF NOT EXISTS field_options (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                field_key TEXT NOT NULL,
                value TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(field_key, value COLLATE NOCASE)
            );
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_field_options_key ON field_options(field_key);")

        conn.execute("""
            CREATE TABLE IF NOT EXISTS admin_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
        """)

        # seed exactly one default admin account, only if no admin users
        # exist yet - never runs again once at least one user exists, so
        # it won't recreate "admin" after someone deletes/renames it
        admin_count = conn.execute("SELECT COUNT(*) AS c FROM admin_users").fetchone()[0]
        if admin_count == 0:
            conn.execute(
                "INSERT INTO admin_users (username, password_hash, created_at) VALUES (?, ?, ?)",
                (DEFAULT_ADMIN_USERNAME, generate_password_hash(DEFAULT_ADMIN_PASSWORD),
                 datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            )

        # seed defaults only if a field_key has no rows yet, so re-running
        # init_db() never overwrites options the user has already edited
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for field_key, options in DEFAULT_OPTIONS.items():
            count = conn.execute(
                "SELECT COUNT(*) AS c FROM field_options WHERE field_key = ?", (field_key,)
            ).fetchone()[0]
            if count == 0:
                for opt in options:
                    conn.execute(
                        "INSERT OR IGNORE INTO field_options (field_key, value, created_at) VALUES (?, ?, ?)",
                        (field_key, opt, now),
                    )

        conn.commit()
    finally:
        conn.close()


def get_options(field_key):
    db = get_db()
    rows = db.execute(
        "SELECT id, value FROM field_options WHERE field_key = ? ORDER BY id ASC",
        (field_key,),
    ).fetchall()
    return [dict(r) for r in rows]


def get_custom_fields():
    db = get_db()
    rows = db.execute(
        "SELECT id, label FROM custom_fields ORDER BY sort_order ASC, id ASC"
    ).fetchall()
    return [dict(r) for r in rows]


def get_default_country_code(options):
    """Pre-select India if present, else fall back to the first option in
    the (Master-Settings-editable) list, else empty if the list is empty."""
    for opt in options:
        if opt["value"] == DEFAULT_COUNTRY_CODE_VALUE:
            return opt["value"]
    return options[0]["value"] if options else ""


def parse_custom_data(raw):
    if not raw:
        return {}
    try:
        data = json.loads(raw)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def get_admin_users():
    db = get_db()
    rows = db.execute(
        "SELECT id, username, created_at FROM admin_users ORDER BY id ASC"
    ).fetchall()
    return [dict(r) for r in rows]


def find_admin_user(username):
    db = get_db()
    row = db.execute("SELECT * FROM admin_users WHERE username = ?", (username,)).fetchone()
    return dict(row) if row else None


def find_admin_user_by_id(user_id):
    db = get_db()
    row = db.execute("SELECT * FROM admin_users WHERE id = ?", (user_id,)).fetchone()
    return dict(row) if row else None


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

def clean(value):
    if value is None:
        return ""
    value = str(value).strip()
    return value[:500]


def join_multi(form, key):
    values = form.getlist(key)
    return ", ".join(clean(v) for v in values if clean(v))


def get_local_ip():
    """Best-effort LAN IP so we can print a friendly URL other devices
    on the same network can use."""
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        return s.getsockname()[0]
    except Exception:
        return "127.0.0.1"
    finally:
        s.close()


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        user_id = session.get("admin_user_id")
        # Require a complete, valid session: is_admin flag, a user id, and
        # that user must still exist in the database. If any part is
        # missing or stale (e.g. an old cookie from before this feature
        # existed, or the account was deleted), clear it and send back to
        # login cleanly - instead of letting later actions fail with a
        # confusing error.
        if not session.get("is_admin") or not user_id or not find_admin_user_by_id(user_id):
            session.clear()
            return redirect(url_for("settings_login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


@app.after_request
def add_no_cache_headers(response):
    """Every page here depends on the login session or live database state
    (Saved Leads, Master Settings, exports, the entry form's dropdowns).
    Without these headers, a browser or an in-between reverse proxy/CDN on
    a server like AWS can cache one visitor's page and serve that same
    cached copy to a completely different visitor - which can make a
    protected page look like it "opened without asking for a password"
    for someone else, even though the login check itself ran correctly."""
    if not request.path.startswith("/static/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, private"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


# ---------------------------------------------------------------------------
# Routes - entry form
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    today = datetime.now().strftime("%Y-%m-%d")
    try:
        db = get_db()
        total = db.execute("SELECT COUNT(*) AS c FROM leads").fetchone()["c"]
    except Exception:
        total = None

    country_code_options = get_options(FIELD_COUNTRY_CODE)
    return render_template(
        "index.html",
        today=today,
        total=total,
        product_options=get_options(FIELD_PRODUCT_TYPES),
        action_options=get_options(FIELD_ACTIONS),
        coordinator_options=get_options(FIELD_COORDINATOR),
        custom_fields=get_custom_fields(),
        country_code_options=country_code_options,
        default_country_code=get_default_country_code(country_code_options),
    )


@app.route("/save", methods=["POST"])
def save_lead():
    try:
        form = request.form
        name = clean(form.get("name"))
        country_code_raw = clean(form.get("country_code"))
        contact_number = clean(form.get("contact_number"))

        errors = {}
        if not name:
            errors["name"] = "Customer name is required."
        if not contact_number:
            errors["contact_number"] = "Contact number is required."
        elif not contact_number.isdigit() or len(contact_number) != 10:
            errors["contact_number"] = "Contact number must be exactly 10 digits."
        if errors:
            return jsonify({"ok": False, "errors": errors}), 400

        # store just the numeric prefix (e.g. "+91") together with the
        # 10-digit number, so the contact column stays a single clean value
        country_code_prefix = country_code_raw.split(" ")[0] if country_code_raw else ""
        contact = f"{country_code_prefix} {contact_number}".strip()

        custom_fields = get_custom_fields()
        custom_data = {}
        entry_date_value = ""
        source_value = ""
        for field in custom_fields:
            value = clean(form.get(f"custom_{field['id']}"))
            label_lower = field["label"].strip().lower()
            if label_lower == "date":
                entry_date_value = value
            elif label_lower == "source":
                source_value = value
            elif value:
                custom_data[field["label"]] = value

        record = {
            "uid": "L" + datetime.now().strftime("%Y%m%d%H%M%S") + "-" + uuid.uuid4().hex[:6],
            "name": name,
            "contact": contact,
            "entry_date": entry_date_value,
            "source": source_value,
            "product_types": join_multi(form, "product_types"),
            "product_description": clean(form.get("product_description")),
            "dimensions": clean(form.get("dimensions")),
            "material_gsm": clean(form.get("material_gsm")),
            "printing_finishing": clean(form.get("printing_finishing")),
            "est_quantity": clean(form.get("est_quantity")),
            "required_by": clean(form.get("required_by")),
            "remarks": clean(form.get("remarks")),
            "reference_carton": clean(form.get("reference_carton")),
            "actions": join_multi(form, "actions"),
            "vpl_coordinator": clean(form.get("vpl_coordinator")),
            "final_remark": clean(form.get("final_remark")),
            "custom_data": json.dumps(custom_data),
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        db = get_db()
        columns = ", ".join(record.keys())
        placeholders = ", ".join("?" for _ in record)
        db.execute(
            f"INSERT INTO leads ({columns}) VALUES ({placeholders})",
            list(record.values()),
        )
        db.commit()

        return jsonify({"ok": True, "uid": record["uid"]})

    except sqlite3.IntegrityError:
        return jsonify({"ok": False, "error": "Duplicate entry, please try again."}), 409
    except Exception:
        app.logger.exception("Failed to save lead")
        return jsonify({"ok": False, "error": "Server error while saving. Please try again."}), 500


# ---------------------------------------------------------------------------
# Routes - view / search leads
# ---------------------------------------------------------------------------

@app.route("/leads")
@login_required
def leads_page():
    return render_template("leads.html")


@app.route("/api/leads")
@login_required
def api_leads():
    try:
        q = clean(request.args.get("q", ""))
        page = max(int(request.args.get("page", 1)), 1)
        per_page = min(max(int(request.args.get("per_page", 25)), 1), 200)
        offset = (page - 1) * per_page

        db = get_db()
        base_query = "FROM leads"
        params = []
        if q:
            base_query += """ WHERE name LIKE ? OR contact LIKE ?
                               OR product_description LIKE ? OR vpl_coordinator LIKE ?"""
            like = f"%{q}%"
            params = [like, like, like, like]

        total = db.execute(f"SELECT COUNT(*) AS c {base_query}", params).fetchone()["c"]
        rows = db.execute(
            f"SELECT * {base_query} ORDER BY id DESC LIMIT ? OFFSET ?",
            params + [per_page, offset],
        ).fetchall()

        leads = [dict(row) for row in rows]
        return jsonify({
            "ok": True, "total": total, "page": page, "per_page": per_page, "leads": leads,
        })
    except Exception:
        app.logger.exception("Failed to fetch leads")
        return jsonify({"ok": False, "error": "Could not load leads."}), 500


@app.route("/api/leads/delete", methods=["POST"])
@login_required
def api_delete_lead():
    try:
        lead_id = request.form.get("id", type=int)
        if not lead_id:
            return jsonify({"ok": False, "error": "Invalid request."}), 400

        db = get_db()
        db.execute("DELETE FROM leads WHERE id = ?", (lead_id,))
        db.commit()

        return jsonify({"ok": True})
    except Exception:
        app.logger.exception("Failed to delete lead")
        return jsonify({"ok": False, "error": "Server error. Please try again."}), 500


# ---------------------------------------------------------------------------
# Routes - export
# ---------------------------------------------------------------------------

def fetch_all_for_export(q=""):
    db = get_db()
    query = "SELECT * FROM leads"
    params = []
    if q:
        query += """ WHERE name LIKE ? OR contact LIKE ?
                      OR product_description LIKE ? OR vpl_coordinator LIKE ?"""
        like = f"%{q}%"
        params = [like, like, like, like]
    query += " ORDER BY id ASC"
    return db.execute(query, params).fetchall()


def build_export_data(rows):
    """Builds headers + row values with any Master-Settings-defined custom
    fields inserted right after Source, so exports always reflect whatever
    custom fields currently exist."""
    custom_fields = get_custom_fields()
    custom_labels = [f["label"] for f in custom_fields]

    headers = EXPORT_HEADERS[:4] + custom_labels + EXPORT_HEADERS[4:]
    prefix_fields = EXPORT_FIELD_ORDER[:4]
    suffix_fields = EXPORT_FIELD_ORDER[4:]

    data_rows = []
    for row in rows:
        custom_dict = parse_custom_data(row["custom_data"])
        prefix_vals = [row[f] if row[f] is not None else "" for f in prefix_fields]
        custom_vals = [custom_dict.get(label, "") for label in custom_labels]
        suffix_vals = [row[f] if row[f] is not None else "" for f in suffix_fields]
        data_rows.append(prefix_vals + custom_vals + suffix_vals)

    return headers, data_rows


@app.route("/export/csv")
@login_required
def export_csv():
    try:
        q = clean(request.args.get("q", ""))
        rows = fetch_all_for_export(q)
        headers, data_rows = build_export_data(rows)

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        for values in data_rows:
            writer.writerow(values)

        output = buffer.getvalue().encode("utf-8-sig")
        filename = f"VPL_Exhibition_Leads_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        return Response(
            output,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception:
        app.logger.exception("CSV export failed")
        flash("CSV export failed. Please try again.", "error")
        return redirect(url_for("leads_page"))


@app.route("/export/xlsx")
@login_required
def export_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        q = clean(request.args.get("q", ""))
        rows = fetch_all_for_export(q)
        headers, data_rows = build_export_data(rows)

        wb = Workbook()
        ws = wb.active
        ws.title = "Exhibition Leads"

        header_fill = PatternFill(start_color="12283F", end_color="12283F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        for row_idx, values in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # sensible default width for every column, wider for known long ones
        default_width = 16
        wide_headers = {"Product Interested In", "Product Description", "Requirement / Remarks", "Action"}
        for i, header in enumerate(headers, start=1):
            width = 24 if header in wide_headers else default_width
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"VPL_Exhibition_Leads_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception:
        app.logger.exception("Excel export failed")
        flash("Excel export failed. Please try again.", "error")
        return redirect(url_for("leads_page"))


# ---------------------------------------------------------------------------
# Routes - Master Settings (password protected)
# ---------------------------------------------------------------------------

@app.route("/settings/login", methods=["GET", "POST"])
def settings_login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = find_admin_user(username)
        if user and check_password_hash(user["password_hash"], password):
            session.permanent = True
            session["is_admin"] = True
            session["admin_user_id"] = user["id"]
            session["admin_username"] = user["username"]
            next_url = request.args.get("next") or url_for("settings_page")
            return redirect(next_url)
        error = "Incorrect username or password."
    return render_template("settings_login.html", error=error)


@app.route("/settings/logout")
def settings_logout():
    session.pop("is_admin", None)
    session.pop("admin_user_id", None)
    session.pop("admin_username", None)
    return render_template("logged_out.html")


@app.route("/settings")
@login_required
def settings_page():
    fields = [
        {"key": FIELD_PRODUCT_TYPES, "label": FIELD_LABELS[FIELD_PRODUCT_TYPES],
         "options": get_options(FIELD_PRODUCT_TYPES)},
        {"key": FIELD_ACTIONS, "label": FIELD_LABELS[FIELD_ACTIONS],
         "options": get_options(FIELD_ACTIONS)},
        {"key": FIELD_COORDINATOR, "label": FIELD_LABELS[FIELD_COORDINATOR],
         "options": get_options(FIELD_COORDINATOR)},
        {"key": FIELD_COUNTRY_CODE, "label": FIELD_LABELS[FIELD_COUNTRY_CODE],
         "options": get_options(FIELD_COUNTRY_CODE)},
    ]
    return render_template(
        "settings.html",
        fields=fields,
        current_username=session.get("admin_username", ""),
        custom_fields=get_custom_fields(),
    )


@app.route("/settings/api/change-password", methods=["POST"])
@login_required
def settings_change_password():
    """Change the password of the currently logged-in user only.
    Does not create a new account and does not touch anyone else's login."""
    try:
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        user = find_admin_user_by_id(session.get("admin_user_id"))
        if not user:
            return jsonify({"ok": False, "error": "Session expired. Please log in again."}), 401
        if not check_password_hash(user["password_hash"], current_password):
            return jsonify({"ok": False, "error": "Current password is incorrect."}), 400
        if not new_password:
            return jsonify({"ok": False, "error": "New password cannot be empty."}), 400
        if len(new_password) < 4:
            return jsonify({"ok": False, "error": "New password must be at least 4 characters."}), 400
        if new_password != confirm_password:
            return jsonify({"ok": False, "error": "New password and confirm password do not match."}), 400

        db = get_db()
        db.execute(
            "UPDATE admin_users SET password_hash = ? WHERE id = ?",
            (generate_password_hash(new_password), user["id"]),
        )
        db.commit()

        return jsonify({"ok": True})
    except Exception:
        app.logger.exception("Failed to change password")
        return jsonify({"ok": False, "error": "Server error. Please try again."}), 500


@app.route("/settings/api/reset-admin", methods=["POST"])
@login_required
def settings_reset_admin():
    """Safety net: resets the default 'admin' account's password back to
    123123 (creating it again if it was ever renamed/removed). Does not
    touch any other accounts."""
    try:
        current_password = request.form.get("current_password", "")
        user = find_admin_user_by_id(session.get("admin_user_id"))
        if not user:
            return jsonify({"ok": False, "error": "Session expired. Please log in again."}), 401
        if not check_password_hash(user["password_hash"], current_password):
            return jsonify({"ok": False, "error": "Current password is incorrect."}), 400

        db = get_db()
        existing = find_admin_user(DEFAULT_ADMIN_USERNAME)
        if existing:
            db.execute(
                "UPDATE admin_users SET password_hash = ? WHERE id = ?",
                (generate_password_hash(DEFAULT_ADMIN_PASSWORD), existing["id"]),
            )
        else:
            db.execute(
                "INSERT INTO admin_users (username, password_hash, created_at) VALUES (?, ?, ?)",
                (DEFAULT_ADMIN_USERNAME, generate_password_hash(DEFAULT_ADMIN_PASSWORD),
                 datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            )
        db.commit()

        return jsonify({"ok": True, "username": DEFAULT_ADMIN_USERNAME})
    except Exception:
        app.logger.exception("Failed to reset admin account")
        return jsonify({"ok": False, "error": "Server error. Please try again."}), 500


VALID_FIELD_KEYS = {FIELD_PRODUCT_TYPES, FIELD_ACTIONS, FIELD_COORDINATOR, FIELD_COUNTRY_CODE}


@app.route("/settings/api/add", methods=["POST"])
@login_required
def settings_add_option():
    try:
        field_key = clean(request.form.get("field_key"))
        value = clean(request.form.get("value"))

        if field_key not in VALID_FIELD_KEYS:
            return jsonify({"ok": False, "error": "Unknown field."}), 400
        if not value:
            return jsonify({"ok": False, "error": "Please enter a value."}), 400
        if len(value) > 120:
            return jsonify({"ok": False, "error": "Value is too long."}), 400

        db = get_db()
        try:
            db.execute(
                "INSERT INTO field_options (field_key, value, created_at) VALUES (?, ?, ?)",
                (field_key, value, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            )
            db.commit()
        except sqlite3.IntegrityError:
            return jsonify({"ok": False, "error": "This option already exists."}), 409

        return jsonify({"ok": True, "options": get_options(field_key)})
    except Exception:
        app.logger.exception("Failed to add option")
        return jsonify({"ok": False, "error": "Server error. Please try again."}), 500


@app.route("/settings/api/delete", methods=["POST"])
@login_required
def settings_delete_option():
    try:
        option_id = request.form.get("id", type=int)
        field_key = clean(request.form.get("field_key"))

        if field_key not in VALID_FIELD_KEYS or not option_id:
            return jsonify({"ok": False, "error": "Invalid request."}), 400

        db = get_db()
        db.execute("DELETE FROM field_options WHERE id = ? AND field_key = ?", (option_id, field_key))
        db.commit()

        return jsonify({"ok": True, "options": get_options(field_key)})
    except Exception:
        app.logger.exception("Failed to delete option")
        return jsonify({"ok": False, "error": "Server error. Please try again."}), 500


@app.route("/settings/api/customfields/add", methods=["POST"])
@login_required
def settings_add_custom_field():
    try:
        label = clean(request.form.get("label"))
        if not label:
            return jsonify({"ok": False, "error": "Please enter a field name."}), 400
        if len(label) > 80:
            return jsonify({"ok": False, "error": "Field name is too long."}), 400

        db = get_db()
        label_lower = label.strip().lower()
        if label_lower == "date":
            order = 0
        elif label_lower == "source":
            order = 1
        else:
            max_order = db.execute("SELECT COALESCE(MAX(sort_order), 1) AS m FROM custom_fields").fetchone()["m"]
            order = max_order + 1

        try:
            db.execute(
                "INSERT INTO custom_fields (label, created_at, sort_order) VALUES (?, ?, ?)",
                (label, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), order),
            )
            db.commit()
        except sqlite3.IntegrityError:
            return jsonify({"ok": False, "error": "A field with this name already exists."}), 409

        return jsonify({"ok": True, "fields": get_custom_fields()})
    except Exception:
        app.logger.exception("Failed to add custom field")
        return jsonify({"ok": False, "error": "Server error. Please try again."}), 500


@app.route("/settings/api/customfields/delete", methods=["POST"])
@login_required
def settings_delete_custom_field():
    try:
        field_id = request.form.get("id", type=int)
        if not field_id:
            return jsonify({"ok": False, "error": "Invalid request."}), 400

        db = get_db()
        db.execute("DELETE FROM custom_fields WHERE id = ?", (field_id,))
        db.commit()

        return jsonify({"ok": True, "fields": get_custom_fields()})
    except Exception:
        app.logger.exception("Failed to delete custom field")
        return jsonify({"ok": False, "error": "Server error. Please try again."}), 500


# ---------------------------------------------------------------------------
# Health check
# ---------------------------------------------------------------------------

@app.route("/health")
def health():
    try:
        db = get_db()
        db.execute("SELECT 1").fetchone()
        return jsonify({"status": "ok"})
    except Exception:
        return jsonify({"status": "error"}), 500


# ---------------------------------------------------------------------------
# Error handlers
# ---------------------------------------------------------------------------

@app.errorhandler(404)
def not_found(e):
    return render_template("error.html", message="Page not found."), 404


@app.errorhandler(500)
def server_error(e):
    return render_template("error.html", message="Something went wrong. Please try again."), 500


init_db()

if __name__ == "__main__":
    local_ip = get_local_ip()
    print("=" * 60)
    print(" VPL Exhibition Enquiry Form is starting...")
    print(f" On this laptop      : http://127.0.0.1:5000")
    print(f" On other devices    : http://{local_ip}:5000")
    print("   (make sure other devices are on the SAME WiFi network,")
    print("    and this laptop's firewall allows incoming port 5000)")
    print("=" * 60)
    app.run(host="0.0.0.0", port=5000, debug=False)
